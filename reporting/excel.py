"""Excel export functionality for schedules."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import Schedule


def export_to_excel(
    schedules: List[Schedule],
    output_path: str | Path,
) -> None:
    """
    Export schedules to a multi-sheet Excel file.

    Args:
        schedules: List of Schedule objects to export
        output_path: Path to save the Excel file

    Each schedule gets its own sheet with:
    - Weekly timetable grid
    - Detailed course list
    - Statistics summary
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add summary sheet
    _create_summary_sheet(wb, schedules)

    # Add sheet for each schedule
    for idx, schedule in enumerate(schedules, 1):
        sheet_name = f"Schedule_{idx}"
        _create_schedule_sheet(wb, schedule, sheet_name)

    # Save workbook
    wb.save(output_path)


def _create_summary_sheet(wb: Workbook, schedules: List[Schedule]) -> None:
    """Create summary sheet with overview of all schedules."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = "📋 Schedule Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')

    # Headers
    headers = ['Schedule', 'Total Courses', 'Total ECTS', 'Conflicts', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for idx, schedule in enumerate(schedules, 1):
        row = idx + 3
        ws.cell(row, 1, f"Schedule #{idx}")
        ws.cell(row, 2, len(schedule.courses))
        ws.cell(row, 3, schedule.total_credits)
        ws.cell(row, 4, schedule.conflict_count)

        # Status
        status_cell = ws.cell(row, 5)
        if schedule.conflict_count == 0:
            status_cell.value = "✅ Perfect"
            status_cell.fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        elif schedule.conflict_count <= 2:
            status_cell.value = "⚠️ Minor Issues"
            status_cell.fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
        else:
            status_cell.value = "❌ Major Issues"
            status_cell.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')

    # Auto-size columns
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_schedule_sheet(wb: Workbook, schedule: Schedule, sheet_name: str) -> None:
    """Create detailed sheet for a single schedule."""
    ws = wb.create_sheet(sheet_name)

    # Title
    ws['A1'] = f"📅 {sheet_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')

    # Statistics
    ws['A2'] = f"Total Courses: {len(schedule.courses)}"
    ws['B2'] = f"Total ECTS: {schedule.total_credits}"
    ws['C2'] = f"Conflicts: {schedule.conflict_count}"

    # Weekly timetable
    _add_weekly_table(ws, schedule, start_row=4)

    # Course list
    course_list_start = 4 + 12  # After timetable + spacing
    _add_course_list(ws, schedule, start_row=course_list_start)


def _add_weekly_table(ws, schedule: Schedule, start_row: int) -> None:
    """Add weekly timetable grid to worksheet."""
    # Headers
    ws.cell(start_row, 1, "Time Slot").font = Font(bold=True)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    for col, day in enumerate(days, 2):
        cell = ws.cell(start_row, col, day)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Build grid
    grid = defaultdict(lambda: defaultdict(str))
    for course in schedule.courses:
        if not course.schedule:
            continue
        for day, slot in course.schedule:
            cell_text = f"{course.main_code}\n{course.teacher or 'TBA'}"
            grid[slot][day] = cell_text

    # Fill grid
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for slot in range(1, 11):  # Slots 1-10
        row = start_row + slot
        
        # Time slot column
        time_cell = ws.cell(row, 1, f"Slot {slot}")
        time_cell.font = Font(bold=True)
        time_cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        time_cell.alignment = Alignment(horizontal='center')
        time_cell.border = thin_border

        # Day columns
        for col, day in enumerate(days, 2):
            cell = ws.cell(row, col, grid[slot].get(day, ''))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            if grid[slot].get(day):
                cell.fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')

    # Set column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Set row heights
    for row in range(start_row + 1, start_row + 11):
        ws.row_dimensions[row].height = 30


def _add_course_list(ws, schedule: Schedule, start_row: int) -> None:
    """Add detailed course list to worksheet."""
    # Section title
    ws.cell(start_row, 1, "📚 Course Details").font = Font(size=12, bold=True)
    start_row += 2

    # Headers
    headers = ['Code', 'Course Name', 'Type', 'ECTS', 'Instructor', 'Schedule']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Course rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, course in enumerate(sorted(schedule.courses, key=lambda c: c.code), 1):
        row = start_row + idx

        # Format schedule
        schedule_str = ""
        if course.schedule:
            day_slots = defaultdict(list)
            for day, slot in course.schedule:
                day_slots[day].append(str(slot))

            schedule_parts = []
            for day, slots in sorted(day_slots.items()):
                schedule_parts.append(f"{day[:3]}: {','.join(slots)}")
            schedule_str = "; ".join(schedule_parts)

        # Add cells
        ws.cell(row, 1, course.code).border = thin_border
        ws.cell(row, 2, course.name).border = thin_border
        ws.cell(row, 3, course.course_type.upper()).border = thin_border
        ws.cell(row, 4, course.ects).border = thin_border
        ws.cell(row, 5, course.teacher or 'TBA').border = thin_border
        ws.cell(row, 6, schedule_str).border = thin_border

        # Alignment
        for col in range(1, 7):
            ws.cell(row, col).alignment = Alignment(vertical='center', wrap_text=True)

        # Alternating row colors
        if idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row, col).fill = PatternFill(
                    start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'
                )

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25


__all__ = ["export_to_excel"]
