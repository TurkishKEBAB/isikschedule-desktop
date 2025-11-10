"""Test suite for reporting modules."""

import pytest
from pathlib import Path
import tempfile
from core.models import Schedule, Course

@pytest.fixture
def sample_course():
    """Create a sample course for testing."""
    return Course(
        code="COMP101-01",
        main_code="COMP101",
        name="Introduction to Programming",
        faculty="Engineering",
        department="Computer Engineering",
        campus="Maslak",
        course_type="lecture",
        ects=6,
        teacher="Dr. John Doe",
        schedule=[("Monday", 1), ("Wednesday", 1)]
    )

@pytest.fixture
def sample_schedule(sample_course):
    """Create a sample schedule for testing."""
    course2 = Course(
        code="MATH101-01",
        main_code="MATH101",
        name="Calculus I",
        faculty="Engineering",
        department="Mathematics",
        campus="Maslak",
        course_type="lecture",
        ects=6,
        teacher="Dr. Jane Smith",
        schedule=[("Tuesday", 2), ("Thursday", 2)]
    )
    
    return Schedule(courses=[sample_course, course2])


class TestPDFExport:
    """Test PDF export functionality."""
    
    def test_pdf_generation(self, sample_schedule):
        """Test that PDF can be generated without errors."""
        from reporting.pdf import save_schedules_as_pdf
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_schedule.pdf"
            save_schedules_as_pdf([sample_schedule], output_path)
            
            assert output_path.exists()
            assert output_path.stat().st_size > 0
    
    def test_empty_schedules_pdf(self):
        """Test PDF generation with empty schedule list."""
        from reporting.pdf import save_schedules_as_pdf
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "empty.pdf"
            save_schedules_as_pdf([], output_path)
            
            assert output_path.exists()


class TestExcelExport:
    """Test Excel export functionality."""
    
    def test_excel_generation(self, sample_schedule):
        """Test that Excel can be generated without errors."""
        from reporting.excel import export_to_excel
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_schedule.xlsx"
            export_to_excel([sample_schedule], output_path)
            
            assert output_path.exists()
            assert output_path.stat().st_size > 0
    
    def test_excel_structure(self, sample_schedule):
        """Test Excel file structure."""
        from reporting.excel import export_to_excel
        from openpyxl import load_workbook
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test.xlsx"
            export_to_excel([sample_schedule], output_path)
            
            wb = load_workbook(output_path)
            
            # Check sheets exist
            assert "Summary" in wb.sheetnames
            assert "Schedule_1" in wb.sheetnames


class TestCharts:
    """Test chart generation functionality."""
    
    def test_summary_chart(self, sample_schedule):
        """Test summary chart generation."""
        from reporting.charts import generate_summary_chart
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "chart.png"
            fig = generate_summary_chart([sample_schedule], output_path)
            
            assert output_path.exists()
            assert fig is not None
    
    def test_algorithm_comparison_chart(self, sample_schedule):
        """Test algorithm comparison chart."""
        from reporting.charts import generate_algorithm_comparison_chart
        
        results = {
            "DFS": [sample_schedule],
            "BFS": [sample_schedule]
        }
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "comparison.png"
            fig = generate_algorithm_comparison_chart(results, output_path)
            
            assert output_path.exists()
            assert fig is not None
